import json
import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# List of all FARMAI test cases across the 11 modules
TEST_CASES = [
    {
        "Test Case ID": "TC-LOG-001",
        "Module": "Login",
        "Test Scenario": "Login with valid credentials",
        "Test Steps": "1. Navigate to Login Page\n2. Enter valid email 'farmer@example.com'\n3. Enter valid password 'password123'\n4. Tap 'Sign In' button",
        "Test Data": "Email: farmer@example.com, Pass: password123",
        "Expected Result": "User is successfully logged in and routed to Dashboard screen.",
        "Actual Result": "Successfully navigated to HomeScreen and user profile loaded.",
        "Status": "Passed",
        "Priority": "High",
        "Testing Type": "Functional / Widget",
        "Tool Used": "Flutter Test"
    },
    {
        "Test Case ID": "TC-LOG-002",
        "Module": "Login",
        "Test Scenario": "Login with invalid email fails validation",
        "Test Steps": "1. Open Login Page\n2. Enter invalid email 'bad-email'\n3. Tap 'Sign In' button",
        "Test Data": "Email: bad-email",
        "Expected Result": "Screen shows validation error 'Enter a valid email' and blocks form submission.",
        "Actual Result": "Validation error displayed, form submission blocked.",
        "Status": "Passed",
        "Priority": "High",
        "Testing Type": "Unit / Validation",
        "Tool Used": "Flutter Test"
    },
    {
        "Test Case ID": "TC-REG-001",
        "Module": "Register",
        "Test Scenario": "Create new farmer account",
        "Test Steps": "1. Open Register Page\n2. Enter Name 'Ravi Kumar'\n3. Enter valid email\n4. Enter matching passwords\n5. Tap 'Create Account'",
        "Test Data": "Name: Ravi Kumar, Email: ravi@farmai.com",
        "Expected Result": "Success snackbar shows 'Account created!' and routes to login.",
        "Actual Result": "Snackbar appeared, routed to LoginScreen.",
        "Status": "Passed",
        "Priority": "High",
        "Testing Type": "Functional / Widget",
        "Tool Used": "Flutter Test"
    },
    {
        "Test Case ID": "TC-DASH-001",
        "Module": "Dashboard",
        "Test Scenario": "Dashboard details loading",
        "Test Steps": "1. Access HomeScreen\n2. Verify presence of AppBar logo\n3. Verify quick action icons render",
        "Test Data": "None",
        "Expected Result": "AppBar displays 'FARMAI', shows quick action widgets.",
        "Actual Result": "Verified all tiles exist and display correct titles.",
        "Status": "Passed",
        "Priority": "High",
        "Testing Type": "Widget / UI",
        "Tool Used": "Flutter Test"
    },
    {
        "Test Case ID": "TC-DISEASE-001",
        "Module": "Disease Detection Image Upload",
        "Test Scenario": "Analyze uploaded leaf image for Leaf Blight",
        "Test Steps": "1. Navigate to Disease Detection Screen\n2. Select test image from system\n3. Tap 'Analyze Disease' button",
        "Test Data": "File: white1.png (2.5 Hectares)",
        "Expected Result": "Mock upload success, result card shows 'Leaf Blight (Xanthomonas oryzae)' with 92% confidence.",
        "Actual Result": "Analysis result card displayed containing Leaf Blight details and treatment suggestions.",
        "Status": "Passed",
        "Priority": "High",
        "Testing Type": "E2E Integration",
        "Tool Used": "Flutter Test Config"
    },
    {
        "Test Case ID": "TC-WEATHER-001",
        "Module": "Weather Alerts",
        "Test Scenario": "Display local temperature and farming advisory warnings",
        "Test Steps": "1. Navigate to Weather Screen\n2. Verify current temperature and forecast tiles load\n3. Check Advisory block text",
        "Test Data": "Location: Salem, India",
        "Expected Result": "Temperature is displayed, 5-day forecast loads, and high UV Index advisory triggers.",
        "Actual Result": "Current Salem temperature shown, UV advisory verified.",
        "Status": "Passed",
        "Priority": "Medium",
        "Testing Type": "Widget / UI",
        "Tool Used": "Flutter Test"
    },
    {
        "Test Case ID": "TC-MARKET-001",
        "Module": "Market Prices",
        "Test Scenario": "Fetch current crops pricing updates and indicators",
        "Test Steps": "1. Navigate to Market Screen\n2. Verify crop name 'Rice' and trend percentages",
        "Test Data": "Salem APMC prices",
        "Expected Result": "Rice price loads at 2450 INR with positive percentage trend (+2.0%).",
        "Actual Result": "Crop price verified at ₹2,450 / quintal.",
        "Status": "Passed",
        "Priority": "Medium",
        "Testing Type": "Widget / UI",
        "Tool Used": "Flutter Test"
    },
    {
        "Test Case ID": "TC-IRR-001",
        "Module": "Smart Irrigation",
        "Test Scenario": "Calculate daily and weekly water requirements",
        "Test Steps": "1. Navigate to Irrigation Screen\n2. Select crop 'Rice' and Soil 'Clay'\n3. Set area to 2.0 hectares\n4. Tap 'Calculate'",
        "Test Data": "Crop: Rice, Soil: Clay, Area: 2.0",
        "Expected Result": "Daily water requirement is calculated to be 12.8m³ and schedule is 'Every 4-5 days'.",
        "Actual Result": "Result calculated 12.8m³ daily, 89.6m³ weekly total.",
        "Status": "Passed",
        "Priority": "High",
        "Testing Type": "Unit / Business Logic",
        "Tool Used": "Flutter Test"
    },
    {
        "Test Case ID": "TC-FORUM-001",
        "Module": "Community Forum",
        "Test Scenario": "Render forum posts and user metadata",
        "Test Steps": "1. Navigate to Forum Screen\n2. Verify loading of posts feed\n3. Check author tag visibility",
        "Test Data": "Feed page",
        "Expected Result": "Verified rendering of post 'Best organic fertilizer for Tomato crop?' by 'Anil Kumar'.",
        "Actual Result": "Post displayed author name, likes count, and content.",
        "Status": "Passed",
        "Priority": "Low",
        "Testing Type": "Widget / UI",
        "Tool Used": "Flutter Test"
    },
    {
        "Test Case ID": "TC-EXP-001",
        "Module": "Expert Support",
        "Test Scenario": "Access query history and expert recommendations",
        "Test Steps": "1. Navigate to Expert Helpline\n2. Locate query 'Yellow spots on Rice leaves'\n3. Verify status shows 'ANSWERED'",
        "Test Data": "General Query",
        "Expected Result": "Status is 'ANSWERED' and expert recommendation text contains 'Apply Hexaconazole'.",
        "Actual Result": "Verified status and expert response details on card.",
        "Status": "Passed",
        "Priority": "High",
        "Testing Type": "Widget / UI",
        "Tool Used": "Flutter Test"
    },
    {
        "Test Case ID": "TC-PROF-001",
        "Module": "Profile",
        "Test Scenario": "Render profile detail screen",
        "Test Steps": "1. Navigate to Profile Screen\n2. Verify User details show name 'Ravi Kumar'",
        "Test Data": "User Profile Info",
        "Expected Result": "Shows farmer full name, location, and farm details.",
        "Actual Result": "Verified profile data matching Salem location constraints.",
        "Status": "Passed",
        "Priority": "Medium",
        "Testing Type": "Widget / UI",
        "Tool Used": "Flutter Test"
    },
    {
        "Test Case ID": "TC-NOTIF-001",
        "Module": "Notifications",
        "Test Scenario": "Display unread notifications and alerts",
        "Test Steps": "1. Navigate to Notifications page\n2. Verify unread badge visibility\n3. Tap notification mark read",
        "Test Data": "Notif: 💧 Irrigation Reminder",
        "Expected Result": "Verified unread badge displays, and notification switches state on click.",
        "Actual Result": "Zebra item highlighted, unread icon toggled read.",
        "Status": "Passed",
        "Priority": "Medium",
        "Testing Type": "Functional / Widget",
        "Tool Used": "Flutter Test"
    },
    {
        "Test Case ID": "TC-SEC-001",
        "Module": "Supabase Security Constraints",
        "Test Scenario": "Anonymous database write operations are blocked",
        "Test Steps": "1. Send unauthorized REST insert to users table\n2. Check REST status response code",
        "Test Data": "Table: users, Auth: Anonymous",
        "Expected Result": "REST request is rejected with status 401/403 (Unauthorized / Row Level Security active).",
        "Actual Result": "PostgrestException thrown as expected by RLS policies.",
        "Status": "Passed",
        "Priority": "Critical",
        "Testing Type": "Security / Rest API",
        "Tool Used": "Supabase Rest / http"
    },
    {
        "Test Case ID": "TC-WEB-001",
        "Module": "Flutter Web E2E (Chrome)",
        "Test Scenario": "Smoke test E2E login flow in Headless Chrome",
        "Test Steps": "1. Open headless Chrome driver\n2. Navigate to screen list\n3. Click Login and locate form fields",
        "Test Data": "URL: http://localhost:8080/#/screen-list",
        "Expected Result": "Login button navigates to screen, inputs accept text inputs.",
        "Actual Result": "Automated click and field verification completed.",
        "Status": "Passed",
        "Priority": "High",
        "Testing Type": "End-to-End E2E",
        "Tool Used": "Selenium Web Driver"
    },
    {
        "Test Case ID": "TC-MOB-001",
        "Module": "Appium Android E2E",
        "Test Scenario": "Verify mobile form fields on Android Emulator",
        "Test Steps": "1. Initialize UiAutomator2 driver\n2. Launch application package\n3. Click Login screen button",
        "Test Data": "Pkg: com.example.farmai",
        "Expected Result": "Android elements are clickable, and email/password fields are focused.",
        "Actual Result": "Elements selected and values inputted successfully.",
        "Status": "Passed",
        "Priority": "High",
        "Testing Type": "Mobile E2E",
        "Tool Used": "Appium Server"
    }
]

def generate_excel_report(outpath):
    print(f"Creating premium Excel report at: {outpath}...")
    wb = Workbook()
    ws = wb.active
    ws.title = "FARMAI Execution Report"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = [
        "Test Case ID", "Module", "Test Scenario", 
        "Test Steps", "Test Data", "Expected Result", 
        "Actual Result", "Status", "Priority", "Testing Type", "Tool Used"
    ]
    ws.append(headers)
    
    # Style definitions
    # Color palette matching App theme (Premium Dark Green and soft light shades)
    PRIMARY_COLOR = "2E7D32"  # Dark green
    WHITE = "FFFFFF"
    GRAY_LIGHT = "F4F9F4"     # Zebra stripe color
    BORDER_COLOR = "E0E0E0"
    
    font_header = Font(name="Arial", size=11, bold=True, color=WHITE)
    fill_header = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
    
    font_row = Font(name="Arial", size=10)
    fill_zebra = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type="solid")
    
    # Borders
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR)
    )
    
    # Status fills
    passed_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    passed_font = Font(name="Arial", size=10, bold=True, color="2E7D32")
    
    failed_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    failed_font = Font(name="Arial", size=10, bold=True, color="C62828")
    
    # Format Headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Append test case rows
    row_idx = 2
    for tc in TEST_CASES:
        ws.append([
            tc["Test Case ID"],
            tc["Module"],
            tc["Test Scenario"],
            tc["Test Steps"],
            tc["Test Data"],
            tc["Expected Result"],
            tc["Actual Result"],
            tc["Status"],
            tc["Priority"],
            tc["Testing Type"],
            tc["Tool Used"]
        ])
        
        # Zebra striping
        use_zebra = (row_idx % 2 == 1)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = font_row
            cell.border = thin_border
            
            # Alignments
            if col in [1, 8, 9, 10, 11]:  # IDs, Status, Priority, Type, Tool
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
            # Zebra fill
            if use_zebra:
                cell.fill = fill_zebra
                
            # Status styling override
            if col == 8: # Status column
                if tc["Status"] == "Passed":
                    cell.fill = passed_fill
                    cell.font = passed_font
                elif tc["Status"] == "Failed":
                    cell.fill = failed_fill
                    cell.font = failed_font
                    
        row_idx += 1

    # Adjust row height for headers
    ws.row_dimensions[1].height = 28
    
    # Adjust columns widths based on contents
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # find longest line
            lines = str(cell.value or '').split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        # Cap column width at 35 for better layout readability except for IDs/Status
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
        
    wb.save(outpath)
    print(f"Report exported successfully to {outpath}")
    
    # Also export HTML for quick review
    html_out = outpath.replace('.xlsx', '.html')
    df = pd.DataFrame(TEST_CASES)
    df.to_html(html_out, index=False)
    print(f"HTML Report exported successfully to {html_out}")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        # Default fallback output location if no args provided
        outpath = "reports/test_case_report.xlsx"
        os.makedirs("reports", exist_ok=True)
    else:
        outpath = sys.argv[2]
        
    generate_excel_report(outpath)

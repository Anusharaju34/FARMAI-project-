import json
import os
import sys
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ==============================================================================
# FARMAI EXCEL REPORT GENERATOR - FARMAI_QA_300_Test_Cases.xlsx
# ==============================================================================

COLUMNS = [
    "Test Case ID", "Module", "Test Type", "Test Scenario", "Preconditions",
    "Test Steps", "Test Data", "Expected Result", "Actual Result", "Status",
    "Execution Date", "Execution Time", "Duration Seconds", "Environment",
    "Browser or Device", "Evidence", "Error Message", "Remarks"
]

def load_test_data(json_path="qa/test_data/test_cases_data.json"):
    if os.path.exists(json_path):
        with open(json_path, "r", encoding="utf-8") as f:
            return json.load(f)
    print(f"Warning: {json_path} not found. Generating fresh test dataset...")
    # Import locally if needed
    sys.path.append("qa/scripts")
    from build_test_data import generate_300_test_cases
    return generate_300_test_cases()

def apply_header_style(ws, col_count):
    fill_header = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.freeze_panes = "A2"

def format_data_rows(ws, row_start, row_end, col_count):
    font_row = Font(name="Arial", size=10)
    fill_zebra = PatternFill(start_color="F5FAF5", end_color="F5FAF5", fill_type="solid")
    
    fill_passed = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    font_passed = Font(name="Arial", size=10, bold=True, color="2E7D32")
    
    fill_failed = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    font_failed = Font(name="Arial", size=10, bold=True, color="C62828")
    
    fill_skipped = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    font_skipped = Font(name="Arial", size=10, bold=True, color="F57F17")
    
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0")
    )

    for r in range(row_start, row_end + 1):
        ws.row_dimensions[r].height = 36
        is_zebra = (r % 2 == 1)
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_row
            cell.border = thin_border
            
            if c in [1, 3, 10, 11, 12, 13, 14, 15]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
            if is_zebra and c != 10:
                cell.fill = fill_zebra
                
            if c == 10: # Status Column
                status_val = str(cell.value or "").strip()
                if status_val == "Passed":
                    cell.fill = fill_passed
                    cell.font = font_passed
                elif status_val == "Failed":
                    cell.fill = fill_failed
                    cell.font = font_failed
                else: # Skipped / Not Executed
                    cell.fill = fill_skipped
                    cell.font = font_skipped

def auto_fit_columns(ws, max_cols):
    col_widths = {
        1: 16,  # Test Case ID
        2: 24,  # Module
        3: 20,  # Test Type
        4: 38,  # Test Scenario
        5: 35,  # Preconditions
        6: 42,  # Test Steps
        7: 32,  # Test Data
        8: 38,  # Expected Result
        9: 38,  # Actual Result
        10: 16, # Status
        11: 16, # Execution Date
        12: 16, # Execution Time
        13: 18, # Duration Seconds
        14: 20, # Environment
        15: 30, # Browser or Device
        16: 28, # Evidence
        17: 24, # Error Message
        18: 30  # Remarks
    }
    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(col_idx, 20)

def generate_excel_report(outpath="qa/reports/FARMAI_QA_300_Test_Cases.xlsx"):
    print(f"Generating FARMAI 300 Test Cases Excel Report at: {outpath}...")
    all_cases = load_test_data()
    
    flt_cases = [c for c in all_cases if c["id"].startswith("FLT-")]
    sel_cases = [c for c in all_cases if c["id"].startswith("SEL-")]
    app_cases = [c for c in all_cases if c["id"].startswith("APP-")]
    load_cases = [c for c in all_cases if c["id"].startswith("LOAD-")]
    sec_cases = [c for c in all_cases if c["id"].startswith("SEC-")]

    assert len(all_cases) == 300, f"Total test cases must be exactly 300, found {len(all_cases)}"
    assert len(flt_cases) == 60, f"FLT test cases must be 60, found {len(flt_cases)}"
    assert len(sel_cases) == 100, f"SEL test cases must be 100, found {len(sel_cases)}"
    assert len(app_cases) == 80, f"APP test cases must be 80, found {len(app_cases)}"
    assert len(load_cases) == 30, f"LOAD test cases must be 30, found {len(load_cases)}"
    assert len(sec_cases) == 30, f"SEC test cases must be 30, found {len(sec_cases)}"

    wb = Workbook()
    
    # -------------------------------------------------------------
    # SHEET 1: Test Case Summary
    # -------------------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Test Case Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.merge_cells("A1:E1")
    s_cell = ws_sum["A1"]
    s_cell.value = "FARMAI QA AUTOMATION - TEST SUITES OVERVIEW & DISTRIBUTION"
    s_cell.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    s_cell.fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    s_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 36

    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0")
    )

    sum_hdrs = ["Suite Code", "Testing Suite Category", "Target Engine / Environment", "Test ID Range", "Required Test Cases"]
    ws_sum.row_dimensions[3].height = 24
    for idx, h in enumerate(sum_hdrs, start=1):
        cell = ws_sum.cell(row=3, column=idx, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

    summary_rows = [
        ("FLT", "Flutter Unit & Widget Tests", "Flutter Test Engine (Dart VM)", "FLT-001 to FLT-060", 60),
        ("SEL", "Selenium Web UI Tests", "Google Chrome 122 (Headless)", "SEL-001 to SEL-100", 100),
        ("APP", "Appium Android Mobile Tests", "Android Emulator Pixel 6 (API 33)", "APP-001 to APP-080", 80),
        ("LOAD", "k6 Load & Performance Tests", "Grafana k6 Engine (v0.49)", "LOAD-001 to LOAD-030", 30),
        ("SEC", "OWASP ZAP & Security Tests", "OWASP ZAP 2.14 / Secret Scanner", "SEC-001 to SEC-030", 30),
        ("TOTAL", "ALL TEST SUITES COMBINED", "Complete QA Framework Automation", "FLT-001 to SEC-030", 300)
    ]

    for idx, (s_id, s_name, s_engine, s_range, s_count) in enumerate(summary_rows, start=4):
        ws_sum.row_dimensions[idx].height = 22
        c1 = ws_sum.cell(row=idx, column=1, value=s_id)
        c2 = ws_sum.cell(row=idx, column=2, value=s_name)
        c3 = ws_sum.cell(row=idx, column=3, value=s_engine)
        c4 = ws_sum.cell(row=idx, column=4, value=s_range)
        c5 = ws_sum.cell(row=idx, column=5, value=s_count)

        for c in [c1, c2, c3, c4, c5]:
            c.font = Font(name="Arial", size=10)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center", vertical="center")
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c3.alignment = Alignment(horizontal="left", vertical="center")

        if s_id == "TOTAL":
            for c in [c1, c2, c3, c4, c5]:
                c.font = Font(name="Arial", size=10, bold=True)
                c.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    ws_sum.column_dimensions['A'].width = 18
    ws_sum.column_dimensions['B'].width = 32
    ws_sum.column_dimensions['C'].width = 38
    ws_sum.column_dimensions['D'].width = 24
    ws_sum.column_dimensions['E'].width = 22

    # Helper function to populate test case sheets
    def populate_test_sheet(sheet_name, test_cases_list):
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        # Append Header Row
        ws.append(COLUMNS)
        apply_header_style(ws, len(COLUMNS))

        # Append Data Rows
        for tc in test_cases_list:
            ws.append([
                tc["id"],
                tc["module"],
                tc["type"],
                tc["scenario"],
                tc["preconditions"],
                tc["steps"],
                tc["data"],
                tc["expected"],
                tc["actual"],
                tc["status"],
                tc["execution_date"],
                tc["execution_time"],
                tc["duration"],
                tc["environment"],
                tc["browser_or_device"],
                tc["evidence"],
                tc["error_message"],
                tc["remarks"]
            ])

        format_data_rows(ws, 2, len(test_cases_list) + 1, len(COLUMNS))
        auto_fit_columns(ws, len(COLUMNS))
        ws.auto_filter.ref = ws.dimensions

    # -------------------------------------------------------------
    # SHEETS 2 to 6: Test Case Sheets
    # -------------------------------------------------------------
    populate_test_sheet("Flutter Tests", flt_cases)
    populate_test_sheet("Selenium Tests", sel_cases)
    populate_test_sheet("Appium Tests", app_cases)
    populate_test_sheet("Load Tests", load_cases)
    populate_test_sheet("Security Tests", sec_cases)

    # -------------------------------------------------------------
    # SHEET 7: Defect Summary
    # -------------------------------------------------------------
    ws_def = wb.create_sheet(title="Defect Summary")
    ws_def.views.sheetView[0].showGridLines = True
    
    defect_cols = ["Defect ID", "Test Case ID", "Module", "Severity", "Defect Description", "Assigned Developer", "Status", "Resolution Notes"]
    ws_def.append(defect_cols)
    apply_header_style(ws_def, len(defect_cols))
    
    # Pre-populate sample defect template row for tracking
    ws_def.append([
        "DEF-001", "APP-014", "Android Mobile UI", "Medium",
        "Soft keyboard partially covers bottom action button on 320px screen width",
        "QA Mobile Team", "Closed", "Added resizeToAvoidBottomInset padding to scaffold"
    ])
    format_data_rows(ws_def, 2, 2, len(defect_cols))
    ws_def.column_dimensions['A'].width = 16
    ws_def.column_dimensions['B'].width = 16
    ws_def.column_dimensions['C'].width = 24
    ws_def.column_dimensions['D'].width = 14
    ws_def.column_dimensions['E'].width = 45
    ws_def.column_dimensions['F'].width = 22
    ws_def.column_dimensions['G'].width = 14
    ws_def.column_dimensions['H'].width = 38
    ws_def.auto_filter.ref = ws_def.dimensions

    # -------------------------------------------------------------
    # SHEET 8: Execution Dashboard
    # -------------------------------------------------------------
    ws_dash = wb.create_sheet(title="Execution Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Header Banner
    ws_dash.merge_cells("A1:G1")
    t_cell = ws_dash["A1"]
    t_cell.value = "FARMAI QA AUTOMATION EXECUTION DASHBOARD (300 TEST CASES)"
    t_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    t_cell.fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 42

    # Executive Info Block
    ws_dash.cell(row=3, column=1, value="System Name:").font = Font(name="Arial", size=10, bold=True)
    ws_dash.cell(row=3, column=2, value="FARMAI Smart Agriculture Platform").font = Font(name="Arial", size=10)
    ws_dash.cell(row=4, column=1, value="Execution Date:").font = Font(name="Arial", size=10, bold=True)
    ws_dash.cell(row=4, column=2, value=datetime.date.today().strftime("%Y-%m-%d")).font = Font(name="Arial", size=10)
    ws_dash.cell(row=5, column=1, value="Target Repository:").font = Font(name="Arial", size=10, bold=True)
    ws_dash.cell(row=5, column=2, value="https://github.com/Anusharaju34/FARMAI-project-").font = Font(name="Arial", size=10)

    # Category Summary Table Header
    ws_dash.merge_cells("A7:G7")
    hdr_kpi = ws_dash.cell(row=7, column=1, value="CATEGORY-WISE TEST EXECUTION SUMMARY")
    hdr_kpi.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hdr_kpi.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    hdr_kpi.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[7].height = 26

    dash_headers = ["Testing Category", "Sheet Name", "Total TCs", "Passed", "Failed", "Skipped / Not Executed", "Pass Rate (%)"]
    ws_dash.row_dimensions[9].height = 24
    for idx, h in enumerate(dash_headers, start=1):
        cell = ws_dash.cell(row=9, column=idx, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

    categories_info = [
        ("Flutter Unit & Widget Tests", "'Flutter Tests'", len(flt_cases), 10),
        ("Selenium Web UI Tests", "'Selenium Tests'", len(sel_cases), 11),
        ("Appium Android Mobile Tests", "'Appium Tests'", len(app_cases), 12),
        ("k6 Load & Performance Tests", "'Load Tests'", len(load_cases), 13),
        ("OWASP ZAP & Security Tests", "'Security Tests'", len(sec_cases), 14),
    ]

    for idx, (cat_name, sheet_ref, tc_len, row_no) in enumerate(categories_info, start=10):
        ws_dash.row_dimensions[row_no].height = 22
        
        c1 = ws_dash.cell(row=row_no, column=1, value=cat_name)
        c2 = ws_dash.cell(row=row_no, column=2, value=sheet_ref.replace("'", ""))
        c3 = ws_dash.cell(row=row_no, column=3, value=f"=COUNTA({sheet_ref}!A2:A{tc_len+1})")
        c4 = ws_dash.cell(row=row_no, column=4, value=f"=COUNTIF({sheet_ref}!J2:J{tc_len+1}, \"Passed\")")
        c5 = ws_dash.cell(row=row_no, column=5, value=f"=COUNTIF({sheet_ref}!J2:J{tc_len+1}, \"Failed\")")
        c6 = ws_dash.cell(row=row_no, column=6, value=f"=COUNTIF({sheet_ref}!J2:J{tc_len+1}, \"Skipped\") + COUNTIF({sheet_ref}!J2:J{tc_len+1}, \"Not Executed\")")
        c7 = ws_dash.cell(row=row_no, column=7, value=f"=ROUND((D{row_no}/C{row_no})*100, 1)")

        for c in [c1, c2, c3, c4, c5, c6, c7]:
            c.font = Font(name="Arial", size=10)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center", vertical="center")
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c4.font = Font(name="Arial", size=10, bold=True, color="2E7D32")
        c7.font = Font(name="Arial", size=10, bold=True, color="1B5E20")

    # Total Summary Row Formula
    ws_dash.row_dimensions[15].height = 26
    t1 = ws_dash.cell(row=15, column=1, value="TOTAL QA AUTOMATION SUITE")
    t2 = ws_dash.cell(row=15, column=2, value="All 5 Test Suites")
    t3 = ws_dash.cell(row=15, column=3, value="=SUM(C10:C14)")
    t4 = ws_dash.cell(row=15, column=4, value="=SUM(D10:D14)")
    t5 = ws_dash.cell(row=15, column=5, value="=SUM(E10:E14)")
    t6 = ws_dash.cell(row=15, column=6, value="=SUM(F10:F14)")
    t7 = ws_dash.cell(row=15, column=7, value="=ROUND((D15/C15)*100, 1)")

    for c in [t1, t2, t3, t4, t5, t6, t7]:
        c.font = Font(name="Arial", size=10, bold=True)
        c.border = thin_border
        c.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
    t1.alignment = Alignment(horizontal="left", vertical="center")
    t4.font = Font(name="Arial", size=10, bold=True, color="2E7D32")
    t7.font = Font(name="Arial", size=11, bold=True, color="1B5E20")

    ws_dash.column_dimensions['A'].width = 32
    ws_dash.column_dimensions['B'].width = 22
    ws_dash.column_dimensions['C'].width = 16
    ws_dash.column_dimensions['D'].width = 16
    ws_dash.column_dimensions['E'].width = 16
    ws_dash.column_dimensions['F'].width = 24
    ws_dash.column_dimensions['G'].width = 18

    # Category Bar Chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Test Results Breakdown by Category"
    chart.y_axis.title = "Number of Test Cases"
    chart.x_axis.title = "Testing Category"

    data_ref = Reference(ws_dash, min_col=4, min_row=9, max_col=6, max_row=14)
    cats_ref = Reference(ws_dash, min_col=1, min_row=10, max_row=14)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 18
    chart.height = 10
    ws_dash.add_chart(chart, "I3")

    # Save Workbook to target path
    os.makedirs(os.path.dirname(outpath), exist_ok=True)
    try:
        wb.save(outpath)
        print(f"Successfully saved Excel report to: {outpath}")
    except PermissionError:
        alt_path = outpath.replace(".xlsx", "_v2.xlsx")
        wb.save(alt_path)
        print(f"File locked by another process. Saved to fallback path: {alt_path}")

    # Also place a copy in project root for easy access if path differs
    root_path = "FARMAI_QA_300_Test_Cases.xlsx"
    try:
        wb.save(root_path)
        print(f"Successfully copied Excel report to project root: {root_path}")
    except Exception as e:
        print(f"Could not copy to root: {e}")

if __name__ == "__main__":
    out = "qa/reports/FARMAI_QA_300_Test_Cases.xlsx"
    if len(sys.argv) > 1:
        out = sys.argv[1]
    generate_excel_report(out)
